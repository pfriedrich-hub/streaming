from openpyxl import *
import tkinter as tk
from pathlib import Path

def create_workbook(file_path=Path.cwd() / 'results' / "gui_get_results_streaming_speech.xlsx"):
    wb = load_workbook(filename=file_path)
    ws = wb['Results']
    ws["A1"] = "Subject ID"
    ws["B1"] = "52.5° Button 1"
    ws["C1"] = "52.5° Button 2"
    ws["D1"] = "52.5° Button 3"
    ws["E1"] = "52.5° Button 4"
    ws["F1"] = "35° Button 1"
    ws["G1"] = "35° Button 2"
    ws["H1"] = "35° Button 3"
    ws["I1"] = "35° Button 4"
    ws["J1"] = "17.5° Button 1"
    ws["K1"] = "17.5° Button 2"
    ws["L1"] = "17.5° Button 3"
    ws["M1"] = "17.5° Button 4"

def write_to_xls(file_path, window):
    wb = load_workbook(filename=file_path)
    ws = wb['Results']
    new_line = ws.max_row + 1
    ws.cell(column=1, row=new_line, value=window.children['!entry'].get())
    ws.cell(column=2, row=new_line, value=window.children['!entry2'].get())
    ws.cell(column=3, row=new_line, value=window.children['!entry3'].get())
    ws.cell(column=4, row=new_line, value=window.children['!entry4'].get())
    ws.cell(column=5, row=new_line, value=window.children['!entry5'].get())
    ws.cell(column=6, row=new_line, value=window.children['!entry6'].get())
    ws.cell(column=7, row=new_line, value=window.children['!entry7'].get())
    ws.cell(column=8, row=new_line, value=window.children['!entry8'].get())
    ws.cell(column=9, row=new_line, value=window.children['!entry9'].get())
    ws.cell(column=10, row=new_line, value=window.children['!entry10'].get())
    ws.cell(column=11, row=new_line, value=window.children['!entry11'].get())
    ws.cell(column=12, row=new_line, value=window.children['!entry12'].get())
    ws.cell(column=13, row=new_line, value=window.children['!entry13'].get())
    wb.save(filename=file_path)
    print('Save successful')

def open_gui(file_path=Path.cwd() / 'results' / 'gui_get_results_streaming_speech.xlsx'):
    # GUI
    window = tk.Tk()
    window.geometry('1000x700')
    window.title('Eingabeparameter')
    window.config(background='#4682B4', padx=10, pady=10)
    # labels for entries
    # labels
    lbl_1 = tk.Label(window, text="Subject ID", fg='black', font=("Helvetica", 16, 'bold'), width=16)  # subject ID
    lbl_1.place(x=40, y=52)
    txtfld_1 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)  # subject ID
    txtfld_1.place(x=200, y=50)
    # condition 1: 52.5°
    lbl_cond_1 = tk.Label(window, text="Distance 52.5°", fg='black', font=("Helvetica", 16, 'bold'), width=16)
    lbl_cond_1.place(x=80, y=150)
    lbl_2 = tk.Label(window, text="Button 1", fg='black', font=("Helvetica", 16), width=10)  # cond 1, button 1
    lbl_2.place(x=40, y=200)
    txtfld_2 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)  # cond 1 button 1
    txtfld_2.place(x=160, y=200)
    lbl_3 = tk.Label(window, text="Button 2", fg='black', font=("Helvetica", 16), width=10)  # cond 1, button 2
    lbl_3.place(x=40, y=250)
    txtfld_3 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)  # cond 1 button 2
    txtfld_3.place(x=160, y=250)
    lbl_4 = tk.Label(window, text="Button 3", fg='black', font=("Helvetica", 16), width=10)  # cond 1, button 3
    lbl_4.place(x=40, y=300)
    txtfld_4 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)  # cond 1 button 3
    txtfld_4.place(x=160, y=300)
    lbl_5 = tk.Label(window, text="Button 4", fg='black', font=("Helvetica", 16), width=10)  # cond 1, button 4
    lbl_5.place(x=40, y=350)
    txtfld_5 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)  # cond 1 button 4
    txtfld_5.place(x=160, y=350)
    # condition 2: 35°
    lbl_cond_2 = tk.Label(window, text="Distance 35°", fg='black', font=("Helvetica", 16, 'bold'), width=16)
    lbl_cond_2.place(x=700, y=150)
    lbl_6 = tk.Label(window, text="Button 1", fg='black', font=["Helvetica", 16], width=10)  # cond 2, button 1
    lbl_6.place(x=350, y=200)
    txtfld_6 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)  # cond 2 button 1
    txtfld_6.place(x=470, y=200)
    lbl_7 = tk.Label(window, text="Button 2", fg='black', font=["Helvetica", 16], width=10)  # cond 2, button 2
    lbl_7.place(x=350, y=250)
    txtfld_7 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)  # cond 2 button 2
    txtfld_7.place(x=470, y=250)
    lbl_8 = tk.Label(window, text="Button 3", fg='black', font=["Helvetica", 16], width=10)  # cond 2, button 3
    lbl_8.place(x=350, y=300)
    txtfld_8 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)  # cond 2 button 3
    txtfld_8.place(x=470, y=300)
    lbl_9 = tk.Label(window, text="Button 4", fg='black', font=["Helvetica", 16], width=10)  # cond 2, button 4
    lbl_9.place(x=350, y=350)
    txtfld_9 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)  # cond 2 button 4
    txtfld_9.place(x=470, y=350)
    # condition 3: 17.5°
    lbl_cond_3 = tk.Label(window, text="Distance 17.5°", fg='black', font=("Helvetica", 16, 'bold'), width=16)
    lbl_cond_3.place(x=390, y=150)
    lbl_10 = tk.Label(window, text="Button 1", fg='black', font=["Helvetica", 16], width=10)  # cond 3, button 1
    lbl_10.place(x=660, y=200)
    txtfld_10 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)  # cond 3 button 1
    txtfld_10.place(x=780, y=200)
    lbl_11 = tk.Label(window, text="Button 2", fg='black', font=["Helvetica", 16], width=10)  # cond 3, button 2
    lbl_11.place(x=660, y=250)
    txtfld_11 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)  # cond 3 button 2
    txtfld_11.place(x=780, y=250)
    lbl_12 = tk.Label(window, text="Button 3", fg='black', font=["Helvetica", 16], width=10)  # cond 3, button 3
    lbl_12.place(x=660, y=300)
    txtfld_12 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)  # cond 3 button 3
    txtfld_12.place(x=780, y=300)
    lbl_13 = tk.Label(window, text="Button 4", fg='black', font=["Helvetica", 16], width=10)  # cond 3, button 4
    lbl_13.place(x=660, y=350)
    txtfld_13 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)  # cond 3 button 4
    txtfld_13.place(x=780, y=350)

    # start experiment
    btn_start = tk.Button(window, text="Start Experiment", fg='black', font=('Helvetica', 16, 'bold'),
                          command=write_to_xls)
    btn_start.place(x=700, y=50)
    # save button
    btn = tk.Button(window, text="Save Entry", fg='black', font=('Helvetica', 16),
                    command=write_to_xls(file_path, window))
    btn.place(x=800, y=50)
    # window.mainloop()
    return window


def add_responses(window, condition, responses):
    c_dict = {'52.5°': [2, 3, 4, 5], '35°': [6, 7, 8, 9], '17.5°': [10, 11, 12, 13]}
    window.children['!entry%i' % c_dict[condition][0]].insert(0, responses[0])
    window.children['!entry%i' % c_dict[condition][1]].insert(0, responses[1])
    window.children['!entry%i' % c_dict[condition][2]].insert(0, responses[2])
    window.children['!entry%i' % c_dict[condition][3]].insert(0, responses[3])

def add_close_button(window):
    btn_exit = tk.Button(window, text="Exit", fg='black', font=('Helvetica', 16), command=window.destroy)
    btn_exit.place(x=700, y=100)



"""

from openpyxl import *
import tkinter as tk
from tkinter import ttk


# 1: subject ID
# 2: 52.5° Button 1
# 3. 52.5° Button 2
# 4. 52.5° Button 3
# 5. 52.5° Button 4
# 6: 35.0° Button 1
# 7: 35.0° Button 2
# 8: 35.0° Button 3
# 9: 35.0° Button 4
# 10: 17.5° Button 1
# 11: 17.5° Button 2
# 12: 17.5° Button 3
# 13: 17.5° Button 4

wb = load_workbook(filename="results/gui_get_results_streaming_speech.xlsx")

ws1 = wb['Results']

ws1["A1"] = "Subject ID"
ws1["B1"] = "52.5° Button 1"
ws1["C1"] = "52.5° Button 2"
ws1["D1"] = "52.5° Button 3"
ws1["E1"] = "52.5° Button 4"
ws1["F1"] = "35° Button 1"
ws1["G1"] = "35° Button 2"
ws1["H1"] = "35° Button 3"
ws1["I1"] = "35° Button 4"
ws1["J1"] = "17.5° Button 1"
ws1["K1"] = "17.5° Button 2"
ws1["L1"] = "17.5° Button 3"
ws1["M1"] = "17.5° Button 4"


def new_entry():
    new_line = ws1.max_row + 1
    ws1.cell(column=1, row=new_line, value=txtfld_1.get())
    ws1.cell(column=2, row=new_line, value=txtfld_2.get())
    ws1.cell(column=3, row=new_line, value=txtfld_3.get())
    ws1.cell(column=4, row=new_line, value=txtfld_4.get())
    ws1.cell(column=5, row=new_line, value=txtfld_5.get())
    ws1.cell(column=6, row=new_line, value=txtfld_6.get())
    ws1.cell(column=7, row=new_line, value=txtfld_7.get())
    ws1.cell(column=8, row=new_line, value=txtfld_8.get())
    ws1.cell(column=9, row=new_line, value=txtfld_9.get())
    ws1.cell(column=10, row=new_line, value=txtfld_10.get())
    ws1.cell(column=11, row=new_line, value=txtfld_11.get())
    ws1.cell(column=12, row=new_line, value=txtfld_12.get())
    ws1.cell(column=13, row=new_line, value=txtfld_13.get())
    wb.save(filename="results/gui_get_results_streaming_speech.xlsx")


def close_window():
    window.destroy()

# GUI
window = tk.Tk()
window.geometry('1000x700')
window.title('Eingabeparameter')
window.config(background='#4682B4', padx=10, pady=10)


# labels
lbl_1 = tk.Label(window, text="Subject ID", fg='black', font=("Helvetica", 16, 'bold'), width=16)   # subject ID
lbl_1.place(x=40, y=52)

txtfld_1 = tk.Entry(window, bg='white',width=10, fg='black', bd=3)                                  # subject ID
txtfld_1.place(x=200, y=50)

# condition 1: 52.5°
lbl_cond_1 = tk.Label(window, text="Distance 52.5°", fg='black', font=("Helvetica", 16, 'bold'), width=16)
lbl_cond_1.place(x=80, y=150)

lbl_2 = tk.Label(window, text="Button 1", fg='black', font=("Helvetica", 16), width=10)         # cond 1, button 1
lbl_2.place(x=40, y=200)
txtfld_2 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)                             # cond 1 button 1
txtfld_2.place(x=160, y=200)

lbl_3 = tk.Label(window, text="Button 2", fg='black', font=("Helvetica", 16), width=10)         # cond 1, button 2
lbl_3.place(x=40, y=250)
txtfld_3 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)                             # cond 1 button 2
txtfld_3.place(x=160, y=250)

lbl_4 = tk.Label(window, text="Button 3", fg='black', font=("Helvetica", 16), width=10)         # cond 1, button 3
lbl_4.place(x=40, y=300)
txtfld_4 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)                             # cond 1 button 3
txtfld_4.place(x=160, y=300)

lbl_5 = tk.Label(window, text="Button 4", fg='black', font=("Helvetica", 16), width=10)         # cond 1, button 4
lbl_5.place(x=40, y=350)
txtfld_5 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)                             # cond 1 button 4
txtfld_5.place(x=160, y=350)

# condition 2: 35°
lbl_cond_2 = tk.Label(window, text="Distance 35°", fg='black', font=("Helvetica", 16, 'bold'), width=16)
lbl_cond_2.place(x=700, y=150)

lbl_6 = tk.Label(window, text="Button 1", fg='black', font=["Helvetica", 16], width=10)         # cond 2, button 1
lbl_6.place(x=350, y=200)
txtfld_6 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)                             # cond 2 button 1
txtfld_6.place(x=470, y=200)

lbl_7 = tk.Label(window, text="Button 2", fg='black', font=["Helvetica", 16], width=10)         # cond 2, button 2
lbl_7.place(x=350, y=250)
txtfld_7 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)                             # cond 2 button 2
txtfld_7.place(x=470, y=250)

lbl_8 = tk.Label(window, text="Button 3", fg='black', font=["Helvetica", 16], width=10)         # cond 2, button 3
lbl_8.place(x=350, y=300)
txtfld_8 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)                             # cond 2 button 3
txtfld_8.place(x=470, y=300)

lbl_9 = tk.Label(window, text="Button 4", fg='black', font=["Helvetica", 16], width=10)         # cond 2, button 4
lbl_9.place(x=350, y=350)
txtfld_9 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)                             # cond 2 button 4
txtfld_9.place(x=470, y=350)

# condition 3: 17.5°
lbl_cond_3 = tk.Label(window, text="Distance 17.5°", fg='black', font=("Helvetica", 16, 'bold'), width=16)
lbl_cond_3.place(x=390, y=150)

lbl_10 = tk.Label(window, text="Button 1", fg='black', font=["Helvetica", 16], width=10)    # cond 3, button 1
lbl_10.place(x=660, y=200)
txtfld_10 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)                        # cond 3 button 1
txtfld_10.place(x=780, y=200)

lbl_11 = tk.Label(window, text="Button 2", fg='black', font=["Helvetica", 16], width=10)    # cond 3, button 2
lbl_11.place(x=660, y=250)
txtfld_11 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)                        # cond 3 button 2
txtfld_11.place(x=780, y=250)

lbl_12 = tk.Label(window, text="Button 3", fg='black', font=["Helvetica", 16], width=10)    # cond 3, button 3
lbl_12.place(x=660, y=300)
txtfld_12 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)                        # cond 3 button 3
txtfld_12.place(x=780, y=300)

lbl_13 = tk.Label(window, text="Button 4", fg='black', font=["Helvetica", 16], width=10)    # cond 3, button 4
lbl_13.place(x=660, y=350)
txtfld_13 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)                        # cond 3 button 4
txtfld_13.place(x=780, y=350)

# start experiment
btn_start = tk.Button(window, text="Start Experiment", fg='black', font=('Helvetica', 16, 'bold'), command = new_entry)
btn_start.place(x=700, y=50)

# save button - schreibt Daten in excel
btn_save = tk.Button(window, text="Save Entry", fg='black', font=('Helvetica', 16, 'bold'), command = new_entry)
btn_save.place(x=40, y=450)

# exit button - close gui - funktioniert nur, wenn man ihn sofort nach 'save entry' drückt
btn_exit = tk.Button(window, text="Exit", fg='black', font=('Helvetica', 16, 'bold'), command = close_window)
btn_exit.place(x=350, y=450)

wb.save(filename="results/gui_get_results_streaming_speech.xlsx")

window.mainloop()

"""