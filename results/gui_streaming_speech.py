
from openpyxl import *
import tkinter as tk
from tkinter import ttk


# 1: subject ID
# 2: condition
# 3. Target speaker
# 4. Masker speaker left
# 5. Masker speaker right
# 6: Button 1
# 7: Button 2
# 8: Button 3
# 9: Button 4

wb = load_workbook(filename="gui_get_results_streaming_speech.xlsx")
# sheets for conditions 17°, 34°, 45°
ws17 = wb['17°']
ws34 = wb['34°']
ws45 = wb['45°']
for ws in [ws17, ws34, ws45]:
    ws["A1"] = "Subject ID"
    ws["B1"] = "Condition"
    ws["C1"] = "Target Speaker"
    ws["D1"] = "Masker Speaker left"
    ws["E1"] = "Masker Speaker right"
    ws["F1"] = "Button 1"
    ws["G1"] = "Button 2"
    ws["H1"] = "Button 3"
    ws["I1"] = "Button 4"

# zurzeit noch als for-Schleife -> wird in alle sheets eingetragen; muss sich aber auf condition beziehen...
# wenn Paul keine Idee hat, alles 3x untereinander für jedes sheet getrennt schreiben
def new_entry():
    for ws in [ws17, ws34, ws45]:
        new_line = ws.max_row + 1
        ws.cell(column=1, row=new_line, value=txtfld_1.get())
        ws.cell(column=2, row=new_line, value=txtfld_2.get())
        ws.cell(column=3, row=new_line, value=txtfld_3.get())
        ws.cell(column=4, row=new_line, value=txtfld_4.get())
        ws.cell(column=5, row=new_line, value=txtfld_5.get())
        ws.cell(column=6, row=new_line, value=txtfld_6.get())
        ws.cell(column=7, row=new_line, value=txtfld_7.get())
        ws.cell(column=8, row=new_line, value=txtfld_8.get())
        ws.cell(column=9, row=new_line, value=txtfld_9.get())
        wb.save(filename="gui_get_results_streaming_speech.xlsx")

def close_window():
    window.destroy()

# GUI
window = tk.Tk()
window.geometry('1000x700')
window.title('Eingabeparameter')
window.config(background='#4682B4', padx=10, pady=10)



# labels for entries
lbl_1 = tk.Label(window, text="Subject ID", fg='black', font=("Helvetica", 16), width=16)
lbl_1.place(x=40, y=50)

lbl_2 = tk.Label(window, text="Condition", fg='black', font=("Helvetica", 16), width=16)
lbl_2.place(x=40, y=100)

lbl_3 = tk.Label(window, text="Target speaker", fg='black', font=("Helvetica", 16), width=16)
lbl_3.place(x=40, y=150)

lbl_4 = tk.Label(window, text="Masker speaker left", fg='black', font=("Helvetica", 16), width=16)
lbl_4.place(x=40, y=200)

lbl_5 = tk.Label(window, text="Masker speaker right", fg='black', font=("Helvetica", 16), width=16)
lbl_5.place(x=40, y=250)

lbl_6 = tk.Label(window, text="Button 1", fg='black', font=["Helvetica", 16], width=10)
lbl_6.place(x=400, y=50)

lbl_7 = tk.Label(window, text="Button 2", fg='black', font=["Helvetica", 16], width=10)
lbl_7.place(x=400, y=100)

lbl_8 = tk.Label(window, text="Button 3", fg='black', font=["Helvetica", 16], width=10)
lbl_8.place(x=400, y=150)

lbl_9 = tk.Label(window, text="Button 4", fg='black', font=["Helvetica", 16], width=10)
lbl_9.place(x=400, y=200)

# entries
txtfld_1 = tk.Entry(window, bg='white',width=10, fg='black', bd=3)          # subject ID
txtfld_1.place(x=200, y=50)

txtfld_2 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)        # condition
txtfld_2.place(x=200, y=100)


# entries - insert or get from streaming_speech
txtfld_3 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)         # target speaker
txtfld_3.place(x=200, y=150)

txtfld_4 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)         # masker speaker left
txtfld_4.place(x=200, y=200)

txtfld_5 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)         # masker speaker right
txtfld_5.place(x=200, y=250)

txtfld_6 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)         # button 1
txtfld_6.place(x=520, y=50)

txtfld_7 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)         # button 2
txtfld_7.place(x=520, y=100)

txtfld_8 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)         # button 3
txtfld_8.place(x=520, y=150)

txtfld_9 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)         # button 4
txtfld_9.place(x=520, y=200)



#save button
btn = tk.Button(window, text="Save Entry", fg='black', font=('Helvetica', 16 ), command = new_entry)
btn.place(x=700, y=50)

# close button
btn_exit = tk.Button(window, text="Exit", fg='black', font=('Helvetica', 16), command = close_window)
btn_exit.place(x=700, y=100)

wb.save(filename="gui_get_results_streaming_speech.xlsx")

window.mainloop()