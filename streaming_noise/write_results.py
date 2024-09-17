from openpyxl import *
from pathlib import Path
# Excel Python Paket importieren
file_path=Path.cwd() / 'results' / "streaming_noise.xlsx"

def create_workbook():
    wb = load_workbook(filename=file_path)
    ws = wb['results']
    ws["A1"] = "Subject ID"
    ws["B1"] = "52.5° hits"
    ws["C1"] = "52.5° missed"
    ws["D1"] = "52.5° false positive"
    ws["E1"] = "35° hits"
    ws["F1"] = "35° missed"
    ws["G1"] = "35° false positive"
    ws["H1"] = "17.5° hits"
    ws["I1"] = "17.5° missed"
    ws["J1"] = "17.5° false positive"
    wb.save(filename=file_path)

def write_subject(subject_id):
    wb = load_workbook(filename=file_path)
    ws = wb['results']
    new_line = ws.max_row + 1
    ws.cell(column=1, row=new_line, value=subject_id)
    wb.save(filename=file_path)


def write_to_xls(subject_id, condition, hits, missed, false_positive):
    col_dict = {'52.5': [2, 3, 4], '35': [5, 6, 7], '17.5': [8, 9, 10]}
    col_idx = col_dict[str(condition)]
    wb = load_workbook(filename=file_path)
    ws = wb['results']
    for row in ws.iter_rows(1):
        if row[0].value == subject_id:
            row_idx = row[0].row
    ws.cell(column=col_idx[0], row=row_idx, value=hits)
    ws.cell(column=col_idx[1], row=row_idx, value=missed)
    ws.cell(column=col_idx[2], row=row_idx, value=false_positive)
    wb.save(filename=file_path)
    print('Save successful')