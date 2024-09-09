from openpyxl import *
import tkinter as tk
from pathlib import Path

def create_workbook(file_path=Path.cwd() / 'results' / "gui_get_results_streaming_speech.xlsx"):
    wb = load_workbook(filename=file_path)
    ws = wb['Results']
    ws["A1"] = "Subject ID"
    ws["B1"] = "52.5° Button 1"
    ws["C1"] = "52.5° Button 2"
    ws["D1"] = "52.5° Button 3"
    ws["E1"] = "52.5° Button 4"
    ws["F1"] = "35° Button 1"
    ws["G1"] = "35° Button 2"
    ws["H1"] = "35° Button 3"
    ws["I1"] = "35° Button 4"
    ws["J1"] = "17.5° Button 1"
    ws["K1"] = "17.5° Button 2"
    ws["L1"] = "17.5° Button 3"
    ws["M1"] = "17.5° Button 4"

def write_to_xls():
    wb = load_workbook(filename=path)
    ws = wb['Results']
    new_line = ws.max_row + 1
    ws.cell(column=1, row=new_line, value=window.children['!entry'].get())
    ws.cell(column=2, row=new_line, value=window.children['!entry2'].get())
    ws.cell(column=3, row=new_line, value=window.children['!entry3'].get())
    ws.cell(column=4, row=new_line, value=window.children['!entry4'].get())
    ws.cell(column=5, row=new_line, value=window.children['!entry5'].get())
    ws.cell(column=6, row=new_line, value=window.children['!entry6'].get())
    ws.cell(column=7, row=new_line, value=window.children['!entry7'].get())
    ws.cell(column=8, row=new_line, value=window.children['!entry8'].get())
    ws.cell(column=9, row=new_line, value=window.children['!entry9'].get())
    ws.cell(column=10, row=new_line, value=window.children['!entry10'].get())
    ws.cell(column=11, row=new_line, value=window.children['!entry11'].get())
    ws.cell(column=12, row=new_line, value=window.children['!entry12'].get())
    ws.cell(column=13, row=new_line, value=window.children['!entry13'].get())
    wb.save(filename=path)
    print('Save successful')

def open_gui(file_path=Path.cwd() / 'results' / 'gui_get_results_streaming_speech.xlsx'):
    global window, path
    path = file_path
    # GUI
    window = tk.Tk()
    window.geometry('1000x700')
    window.title('Eingabeparameter')
    window.config(background='#4682B4', padx=10, pady=10)
    # labels for entries
    # labels
    lbl_1 = tk.Label(window, text="Subject ID", fg='black', font=("Helvetica", 16, 'bold'), width=16)  # subject ID
    lbl_1.place(x=40, y=52)
    txtfld_1 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)  # subject ID
    txtfld_1.place(x=200, y=50)
    # condition 1: 52.5°
    lbl_cond_1 = tk.Label(window, text="Distance 52.5°", fg='black', font=("Helvetica", 16, 'bold'), width=16)
    lbl_cond_1.place(x=80, y=150)
    lbl_2 = tk.Label(window, text="Button 1", fg='black', font=("Helvetica", 16), width=10)  # cond 1, button 1
    lbl_2.place(x=40, y=200)
    txtfld_2 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)  # cond 1 button 1
    txtfld_2.place(x=160, y=200)
    lbl_3 = tk.Label(window, text="Button 2", fg='black', font=("Helvetica", 16), width=10)  # cond 1, button 2
    lbl_3.place(x=40, y=250)
    txtfld_3 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)  # cond 1 button 2
    txtfld_3.place(x=160, y=250)
    lbl_4 = tk.Label(window, text="Button 3", fg='black', font=("Helvetica", 16), width=10)  # cond 1, button 3
    lbl_4.place(x=40, y=300)
    txtfld_4 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)  # cond 1 button 3
    txtfld_4.place(x=160, y=300)
    lbl_5 = tk.Label(window, text="Button 4", fg='black', font=("Helvetica", 16), width=10)  # cond 1, button 4
    lbl_5.place(x=40, y=350)
    txtfld_5 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)  # cond 1 button 4
    txtfld_5.place(x=160, y=350)
    # condition 2: 35°
    lbl_cond_2 = tk.Label(window, text="Distance 35°", fg='black', font=("Helvetica", 16, 'bold'), width=16)
    lbl_cond_2.place(x=700, y=150)
    lbl_6 = tk.Label(window, text="Button 1", fg='black', font=["Helvetica", 16], width=10)  # cond 2, button 1
    lbl_6.place(x=350, y=200)
    txtfld_6 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)  # cond 2 button 1
    txtfld_6.place(x=470, y=200)
    lbl_7 = tk.Label(window, text="Button 2", fg='black', font=["Helvetica", 16], width=10)  # cond 2, button 2
    lbl_7.place(x=350, y=250)
    txtfld_7 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)  # cond 2 button 2
    txtfld_7.place(x=470, y=250)
    lbl_8 = tk.Label(window, text="Button 3", fg='black', font=["Helvetica", 16], width=10)  # cond 2, button 3
    lbl_8.place(x=350, y=300)
    txtfld_8 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)  # cond 2 button 3
    txtfld_8.place(x=470, y=300)
    lbl_9 = tk.Label(window, text="Button 4", fg='black', font=["Helvetica", 16], width=10)  # cond 2, button 4
    lbl_9.place(x=350, y=350)
    txtfld_9 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)  # cond 2 button 4
    txtfld_9.place(x=470, y=350)
    # condition 3: 17.5°
    lbl_cond_3 = tk.Label(window, text="Distance 17.5°", fg='black', font=("Helvetica", 16, 'bold'), width=16)
    lbl_cond_3.place(x=390, y=150)
    lbl_10 = tk.Label(window, text="Button 1", fg='black', font=["Helvetica", 16], width=10)  # cond 3, button 1
    lbl_10.place(x=660, y=200)
    txtfld_10 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)  # cond 3 button 1
    txtfld_10.place(x=780, y=200)
    lbl_11 = tk.Label(window, text="Button 2", fg='black', font=["Helvetica", 16], width=10)  # cond 3, button 2
    lbl_11.place(x=660, y=250)
    txtfld_11 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)  # cond 3 button 2
    txtfld_11.place(x=780, y=250)
    lbl_12 = tk.Label(window, text="Button 3", fg='black', font=["Helvetica", 16], width=10)  # cond 3, button 3
    lbl_12.place(x=660, y=300)
    txtfld_12 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)  # cond 3 button 3
    txtfld_12.place(x=780, y=300)
    lbl_13 = tk.Label(window, text="Button 4", fg='black', font=["Helvetica", 16], width=10)  # cond 3, button 4
    lbl_13.place(x=660, y=350)
    txtfld_13 = tk.Entry(window, bg='white', width=10, fg='black', bd=3)  # cond 3 button 4
    txtfld_13.place(x=780, y=350)

    # save button
    btn = tk.Button(window, text="Save Entry", fg='black', font=('Helvetica', 16), command=write_to_xls)
    btn.place(x=800, y=50)
    btn_exit = tk.Button(window, text="Exit", fg='black', font=('Helvetica', 16), command=window.destroy)
    btn_exit.place(x=700, y=100)

    # start experiment
    var = tk.IntVar()
    btn_start = tk.Button(window, text="Start Experiment", fg='black', font=('Helvetica', 16, 'bold'),
                          command=lambda: var.set(1))
    btn_start.place(x=700, y=50)
    btn_start.wait_variable(var)       # wait until start button is pressed
    # window.mainloop()
    return window

def add_responses(window, condition, responses):
    c_dict = {'52.5°': [2, 3, 4, 5], '35°': [6, 7, 8, 9], '17.5°': [10, 11, 12, 13]}
    window.children['!entry%i' % c_dict[condition][0]].insert(0, responses[0])
    window.children['!entry%i' % c_dict[condition][1]].insert(0, responses[1])
    window.children['!entry%i' % c_dict[condition][2]].insert(0, responses[2])
    window.children['!entry%i' % c_dict[condition][3]].insert(0, responses[3])

def add_close_button(window):
    btn_exit = tk.Button(window, text="Exit", fg='black', font=('Helvetica', 16), command=window.destroy)
    btn_exit.place(x=700, y=100)

def update():
    tk.update()
